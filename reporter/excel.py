import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DATE_FORMAT = "YYYY-MM-DD HH:MM:SS"
NUMBER_FORMAT = "#,##0.00"


def format_header(ws, num_columns):
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, df):
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for val in df[col_name]:
            val_len = len(str(val)) if val is not None else 0
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)


def apply_column_formats(ws, df):
    for col_idx, col_name in enumerate(df.columns, 1):
        sample = df[col_name].dropna().head(10)
        if sample.empty:
            continue

        is_date = any(isinstance(v, datetime.datetime) for v in sample)
        is_number = any(isinstance(v, (int, float)) for v in sample) and not is_date

        if is_date:
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=col_idx).number_format = DATE_FORMAT
        elif is_number:
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=col_idx).number_format = NUMBER_FORMAT


def generate_report(dataframes, output_path):
    wb = Workbook()
    first = True

    for sheet_name, df in dataframes.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        format_header(ws, len(df.columns))
        auto_width(ws, df)
        apply_column_formats(ws, df)
        ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Reporte generado: {output_path}")
